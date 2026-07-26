import csv
from collections import OrderedDict

from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone

from apps.acolhimento.models import PrimeiroContato


def _responsavel_nome(pessoa):
	responsavel = pessoa.responsavel_atual
	if not responsavel:
		return ''
	return responsavel.get_full_name() or responsavel.get_username()


def _data_br(valor):
	return valor.strftime('%d/%m/%Y') if valor else ''


# Registro canonico de colunas: chave -> (rotulo, funcao de leitura).
# A ordem aqui define a ordem das colunas no arquivo exportado.
COLUNAS_RELATORIO = OrderedDict([
	('nome', ('Nome', lambda p: p.nome)),
	('telefone_whatsapp', ('WhatsApp', lambda p: p.telefone_whatsapp)),
	('email', ('E-mail', lambda p: p.email)),
	('status', ('Status', lambda p: p.get_status_display())),
	('origem_cadastro', ('Origem', lambda p: p.get_origem_cadastro_display())),
	('primeira_vez', ('Primeira vez', lambda p: 'Sim' if p.primeira_vez else 'Nao')),
	('iniciou_interacao', ('Iniciou interacao', lambda p: 'Sim' if p.iniciou_interacao else 'Nao')),
	('como_conheceu', ('Como conheceu', lambda p: p.get_como_conheceu_display())),
	('o_que_busca', ('O que busca', lambda p: p.get_o_que_busca_display())),
	('responsavel', ('Responsavel', _responsavel_nome)),
	('genero', ('Genero', lambda p: p.get_genero_display())),
	('idade', ('Idade', lambda p: p.idade if p.idade is not None else '')),
	('estado_civil', ('Estado civil', lambda p: p.get_estado_civil_display())),
	('religiao', ('Religiao', lambda p: p.religiao)),
	('cidade', ('Cidade', lambda p: p.cidade)),
	('data_primeiro_contato', ('Data 1o contato', lambda p: _data_br(p.data_primeiro_contato))),
	('observacoes', ('Observacoes', lambda p: p.observacoes)),
])

COLUNAS_PADRAO = ['nome', 'telefone_whatsapp', 'status', 'origem_cadastro', 'data_primeiro_contato']

FORMATOS = ['pdf', 'xlsx', 'csv']

TRIESTADO_CHOICES = [('', 'Todos'), ('sim', 'Sim'), ('nao', 'Nao')]


def colunas_choices():
	return [(chave, rotulo) for chave, (rotulo, _) in COLUNAS_RELATORIO.items()]


def normalizar_colunas(selecionadas):
	"""Mantem apenas colunas validas, na ordem canonica de COLUNAS_RELATORIO."""
	selecionadas = set(selecionadas or [])
	return [chave for chave in COLUNAS_RELATORIO if chave in selecionadas]


def filtrar_pessoas(dados):
	queryset = PrimeiroContato.objects.select_related('responsavel_atual').all()

	busca = (dados.get('q') or '').strip()
	if busca:
		queryset = queryset.filter(
			Q(nome__icontains=busca)
			| Q(telefone_whatsapp__icontains=busca)
			| Q(email__icontains=busca)
			| Q(cidade__icontains=busca)
			| Q(religiao__icontains=busca)
		)

	if dados.get('status'):
		queryset = queryset.filter(status=dados['status'])

	if dados.get('origem'):
		queryset = queryset.filter(origem_cadastro=dados['origem'])

	primeira_vez = dados.get('primeira_vez')
	if primeira_vez == 'sim':
		queryset = queryset.filter(primeira_vez=True)
	elif primeira_vez == 'nao':
		queryset = queryset.filter(primeira_vez=False)

	iniciou = dados.get('iniciou_interacao')
	if iniciou == 'sim':
		queryset = queryset.filter(iniciou_interacao=True)
	elif iniciou == 'nao':
		queryset = queryset.filter(iniciou_interacao=False)

	if dados.get('data_inicio'):
		queryset = queryset.filter(data_primeiro_contato__gte=dados['data_inicio'])
	if dados.get('data_fim'):
		queryset = queryset.filter(data_primeiro_contato__lte=dados['data_fim'])

	responsavel = dados.get('responsavel')
	if responsavel == 'sem':
		queryset = queryset.filter(responsavel_atual__isnull=True)
	elif responsavel:
		queryset = queryset.filter(responsavel_atual_id=responsavel)

	return queryset.order_by('nome')


def responsavel_choices():
	from django.contrib.auth import get_user_model

	ids = (
		PrimeiroContato.objects.exclude(responsavel_atual__isnull=True)
		.values_list('responsavel_atual', flat=True)
		.distinct()
	)
	usuarios = get_user_model().objects.filter(id__in=list(ids)).order_by('first_name', 'username')

	opcoes = [('', 'Todos'), ('sem', 'Sem responsavel')]
	for usuario in usuarios:
		opcoes.append((str(usuario.id), usuario.get_full_name() or usuario.get_username()))
	return opcoes


def resumo_pessoas(queryset):
	"""Total e contagem por status (na ordem evolutiva) do conjunto filtrado."""
	from django.db.models import Count

	contagem = {
		linha['status']: linha['n']
		for linha in queryset.values('status').annotate(n=Count('id'))
	}
	por_status = [
		{'rotulo': rotulo, 'total': contagem.get(valor, 0)}
		for valor, rotulo in PrimeiroContato.StatusAcolhimento.choices
		if contagem.get(valor, 0)
	]
	return {'total': sum(contagem.values()), 'por_status': por_status}


def resumo_filtros(dados):
	"""Monta um texto legivel dos filtros aplicados (para o cabecalho do PDF)."""
	partes = []
	if (dados.get('q') or '').strip():
		partes.append(f'Busca "{dados["q"].strip()}"')
	if dados.get('status'):
		partes.append(f'Status: {dict(PrimeiroContato.StatusAcolhimento.choices).get(dados["status"], dados["status"])}')
	if dados.get('origem'):
		partes.append(f'Origem: {dict(PrimeiroContato.OrigemCadastroChoices.choices).get(dados["origem"], dados["origem"])}')
	if dados.get('primeira_vez'):
		partes.append(f'Primeira vez: {dados["primeira_vez"]}')
	if dados.get('iniciou_interacao'):
		partes.append(f'Iniciou interacao: {dados["iniciou_interacao"]}')
	if dados.get('data_inicio'):
		partes.append(f'De {_data_br(dados["data_inicio"])}')
	if dados.get('data_fim'):
		partes.append(f'Ate {_data_br(dados["data_fim"])}')
	responsavel = dados.get('responsavel')
	if responsavel == 'sem':
		partes.append('Sem responsavel')
	elif responsavel:
		from django.contrib.auth import get_user_model
		usuario = get_user_model().objects.filter(id=responsavel).first()
		if usuario:
			partes.append(f'Responsavel: {usuario.get_full_name() or usuario.get_username()}')
	return ' | '.join(partes)


def _nome_arquivo(extensao):
	return f'relatorio_pessoas_{timezone.localtime():%Y%m%d_%H%M}.{extensao}'


def _matriz(pessoas, colunas):
	cabecalho = [COLUNAS_RELATORIO[chave][0] for chave in colunas]
	linhas = [[COLUNAS_RELATORIO[chave][1](pessoa) for chave in colunas] for pessoa in pessoas]
	return cabecalho, linhas


def gerar_csv(pessoas, colunas):
	cabecalho, linhas = _matriz(pessoas, colunas)
	response = HttpResponse(content_type='text/csv; charset=utf-8')
	response['Content-Disposition'] = f'attachment; filename="{_nome_arquivo("csv")}"'
	response.write('﻿')  # BOM: garante acentos corretos ao abrir no Excel
	writer = csv.writer(response)
	writer.writerow(cabecalho)
	for linha in linhas:
		writer.writerow(['' if valor is None else valor for valor in linha])
	return response


def gerar_xlsx(pessoas, colunas):
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, PatternFill
	from openpyxl.utils import get_column_letter

	cabecalho, linhas = _matriz(pessoas, colunas)

	wb = Workbook()
	ws = wb.active
	ws.title = 'Pessoas'
	ws.append(cabecalho)

	header_fill = PatternFill('solid', fgColor='276749')
	header_font = Font(bold=True, color='FFFFFF')
	for cell in ws[1]:
		cell.fill = header_fill
		cell.font = header_font
		cell.alignment = Alignment(vertical='center')

	for linha in linhas:
		ws.append(['' if valor is None else valor for valor in linha])

	ws.freeze_panes = 'A2'
	ws.auto_filter.ref = f'A1:{get_column_letter(len(colunas))}1'

	for indice, chave in enumerate(colunas, start=1):
		largura = len(str(COLUNAS_RELATORIO[chave][0]))
		for linha in linhas:
			largura = max(largura, len(str(linha[indice - 1] if linha[indice - 1] is not None else '')))
		ws.column_dimensions[get_column_letter(indice)].width = min(max(largura + 2, 10), 55)

	response = HttpResponse(
		content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
	)
	response['Content-Disposition'] = f'attachment; filename="{_nome_arquivo("xlsx")}"'
	wb.save(response)
	return response


def _pdf_safe(texto):
	# As fontes core do fpdf2 sao latin-1 (cobre acentos do portugues);
	# troca qualquer caractere fora disso para nao quebrar a geracao.
	return str('' if texto is None else texto).encode('latin-1', 'replace').decode('latin-1')


def gerar_pdf(pessoas, colunas, resumo=''):
	from fpdf import FPDF
	from fpdf.enums import XPos, YPos
	from fpdf.fonts import FontFace

	cabecalho, linhas = _matriz(pessoas, colunas)

	pdf = FPDF(orientation='L', unit='mm', format='A4')
	pdf.set_auto_page_break(auto=True, margin=12)
	pdf.add_page()

	pdf.set_font('Helvetica', 'B', 15)
	pdf.set_text_color(31, 81, 58)
	pdf.cell(0, 9, _pdf_safe('Relatorio de pessoas - Acolhimento PIBVP'),
	         new_x=XPos.LMARGIN, new_y=YPos.NEXT)

	pdf.set_font('Helvetica', '', 9)
	pdf.set_text_color(90, 90, 90)
	pdf.cell(0, 5, _pdf_safe(f'Gerado em {timezone.localtime():%d/%m/%Y %H:%M} - {len(pessoas)} registro(s)'),
	         new_x=XPos.LMARGIN, new_y=YPos.NEXT)
	if resumo:
		pdf.multi_cell(0, 5, _pdf_safe(f'Filtros: {resumo}'),
		               new_x=XPos.LMARGIN, new_y=YPos.NEXT)

	from collections import Counter

	contagem = Counter(pessoa.status for pessoa in pessoas)
	partes_status = [
		f'{rotulo}: {contagem[valor]}'
		for valor, rotulo in PrimeiroContato.StatusAcolhimento.choices
		if contagem.get(valor)
	]
	if partes_status:
		pdf.multi_cell(0, 5, _pdf_safe('Por status: ' + '   '.join(partes_status)),
		               new_x=XPos.LMARGIN, new_y=YPos.NEXT)
	pdf.ln(2)

	pdf.set_text_color(0, 0, 0)
	pdf.set_font('Helvetica', '', 7)
	estilo_cabecalho = FontFace(emphasis='BOLD', color=(255, 255, 255), fill_color=(39, 103, 73))

	with pdf.table(
		headings_style=estilo_cabecalho,
		line_height=4.5,
		text_align='LEFT',
		cell_fill_color=(244, 248, 243),
		cell_fill_mode='ROWS',
	) as table:
		linha_cabecalho = table.row()
		for titulo in cabecalho:
			linha_cabecalho.cell(_pdf_safe(titulo))
		for linha in linhas:
			linha_tabela = table.row()
			for valor in linha:
				linha_tabela.cell(_pdf_safe(valor))

	response = HttpResponse(bytes(pdf.output()), content_type='application/pdf')
	response['Content-Disposition'] = f'attachment; filename="{_nome_arquivo("pdf")}"'
	return response


def gerar_relatorio(queryset, colunas, formato, resumo=''):
	pessoas = list(queryset)
	if formato == 'csv':
		return gerar_csv(pessoas, colunas)
	if formato == 'xlsx':
		return gerar_xlsx(pessoas, colunas)
	return gerar_pdf(pessoas, colunas, resumo)
